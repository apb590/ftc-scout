import openpyxl

wb = openpyxl.load_workbook("/Users/alden/.gemini/antigravity/brain/681cf4a1-18d2-4102-90f6-590248cea9b8/scratch/workbook.xlsx")
sheet = wb["Configuration"]
print("Last row:", sheet.max_row)
print("Last col:", sheet.max_column)

# Print all rows
for r in range(1, sheet.max_row + 1):
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    print(f"Row {r}: {row_vals}")
