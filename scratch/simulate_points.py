import openpyxl

wb = openpyxl.load_workbook("/Users/alden/.gemini/antigravity/brain/681cf4a1-18d2-4102-90f6-590248cea9b8/scratch/workbook.xlsx")
sheet = wb["ScoutData"]
headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]

team_col_idx = headers.index("teamno")
match_col_idx = headers.index("matchno")
preload_col_idx = headers.index("preload_made")
pickup_col_idx = headers.index("pickup_made")
close_col_idx = headers.index("close_made")
far_col_idx = headers.index("far_made")
auto_park_col_idx = headers.index("auto_park")
automove_col_idx = headers.index("automove")
park_base_col_idx = headers.index("park_base")
park_bonus_col_idx = headers.index("park_bonus")

def clean_int(val):
    if val is None:
        return 0
    try:
        return int(float(val))
    except:
        return 0

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip().lower()

matching_rows = []
for r in range(2, sheet.max_row + 1):
    val = sheet.cell(r, team_col_idx + 1).value
    if val in [8393, 17895, "8393", "17895"]:
        matching_rows.append(r)

print(f"Simulation of Dynamic Point Repair for {len(matching_rows)} rows:")
for r in matching_rows:
    team = clean_int(sheet.cell(r, team_col_idx + 1).value)
    match = clean_int(sheet.cell(r, match_col_idx + 1).value)
    preload = clean_int(sheet.cell(r, preload_col_idx + 1).value)
    pickup = clean_int(sheet.cell(r, pickup_col_idx + 1).value)
    close = clean_int(sheet.cell(r, close_col_idx + 1).value)
    far = clean_int(sheet.cell(r, far_col_idx + 1).value)
    
    auto_park = clean_str(sheet.cell(r, auto_park_col_idx + 1).value)
    automove = clean_str(sheet.cell(r, automove_col_idx + 1).value)
    park_base = clean_str(sheet.cell(r, park_base_col_idx + 1).value)
    park_bonus = clean_str(sheet.cell(r, park_bonus_col_idx + 1).value)
    
    preload_pts = preload * 5
    pickup_pts = pickup * 5
    auto_park_pts = 10 if "inside" in auto_park else 0
    auto_move_pts = 5 if automove == "yes" else 0
    calculated_auto = preload_pts + pickup_pts + auto_park_pts + auto_move_pts
    
    close_pts = close * 2
    far_pts = far * 5
    park_base_pts = 10 if "parked" in park_base else 0
    park_bonus_pts = 20 if "parked" in park_bonus else 0
    calculated_tele = close_pts + far_pts + park_base_pts + park_bonus_pts
    
    calculated_total = calculated_auto + calculated_tele
    print(f"Row {r} (Team {team}, Match {match}): Auto={calculated_auto} (preload={preload}, pickup={pickup}, park={auto_park_pts}, move={auto_move_pts}), Tele={calculated_tele} (close={close}, far={far}, park_base={park_base_pts}, park_bonus={park_bonus_pts}), Total={calculated_total}")
