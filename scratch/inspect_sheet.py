import openpyxl

xlsx_path = "/Users/alden/.gemini/antigravity/brain/681cf4a1-18d2-4102-90f6-590248cea9b8/scratch/workbook.xlsx"

try:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb["ScoutData"]
    max_row = sheet.max_row
    
    repaired_count = 0
    sample_rows = []
    
    for r in range(2, max_row + 1):
        # Read action columns
        automove = str(sheet.cell(r, 7).value or "").strip().lower() # Col 7 (Index 6)
        preload_made = sheet.cell(r, 8).value # Col 8 (Index 7)
        pickup_made = sheet.cell(r, 10).value # Col 10 (Index 9)
        auto_park = str(sheet.cell(r, 17).value or "").strip().lower() # Col 17 (Index 16)
        
        close_made = sheet.cell(r, 20).value # Col 20 (Index 19)
        far_made = sheet.cell(r, 23).value # Col 23 (Index 22)
        park_base = str(sheet.cell(r, 32).value or "").strip().lower() # Col 32 (Index 31)
        park_bonus = str(sheet.cell(r, 33).value or "").strip().lower() # Col 33 (Index 32)
        
        # Parse integers
        try:
            preload = int(float(preload_made)) if preload_made is not None and preload_made != "" else 0
        except: preload = 0
            
        try:
            pickup = int(float(pickup_made)) if pickup_made is not None and pickup_made != "" else 0
        except: pickup = 0
            
        try:
            close = int(float(close_made)) if close_made is not None and close_made != "" else 0
        except: close = 0
            
        try:
            far = int(float(far_made)) if far_made is not None and far_made != "" else 0
        except: far = 0
            
        # Point math
        preload_pts = preload * 5
        pickup_pts = pickup * 5
        auto_park_pts = 10 if "inside" in auto_park else 0
        auto_move_pts = 5 if automove == "yes" else 0
        auto_total = preload_pts + pickup_pts + auto_park_pts + auto_move_pts
        
        close_pts = close * 2
        far_pts = far * 5
        base_park_pts = 10 if "parked" in park_base else 0
        bonus_park_pts = 20 if "parked" in park_bonus else 0
        tele_total = close_pts + far_pts + base_park_pts + bonus_park_pts
        
        total_pts = auto_total + tele_total
        
        repaired_count += 1
        
        # Save a sample of Team 8393
        team = sheet.cell(r, 3).value
        try:
            t_num = int(float(team))
        except:
            t_num = None
            
        if t_num in {8393, 17895} and len(sample_rows) < 5:
            sample_rows.append({
                "row": r,
                "team": t_num,
                "match": sheet.cell(r, 4).value,
                "automove": automove,
                "preload": preload,
                "pickup": pickup,
                "close": close,
                "far": far,
                "auto_calc": auto_total,
                "tele_calc": tele_total,
                "total_calc": total_pts
            })
            
    print(f"Total rows simulated for repair: {repaired_count}")
    print("\nSample Repaired Rows for Target Teams:")
    for s in sample_rows:
        print(f"Row {s['row']} | Team {s['team']} | Match {s['match']}")
        print(f"  * Actions: AutoMove={s['automove']}, PreloadMade={s['preload']}, PickupMade={s['pickup']}, CloseMade={s['close']}, FarMade={s['far']}")
        print(f"  * Calculated: AutoPoints={s['auto_calc']}, TeleOpPoints={s['tele_calc']}, TotalPoints={s['total_calc']}")
        
except Exception as e:
    print("Error:", e)
