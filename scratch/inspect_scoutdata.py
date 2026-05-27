import openpyxl

wb = openpyxl.load_workbook("/Users/alden/.gemini/antigravity/brain/681cf4a1-18d2-4102-90f6-590248cea9b8/scratch/workbook.xlsx")
sheet = wb["ScoutData"]
headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]

for r in [5, 7]:
    print(f"\n--- ROW {r} ---")
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    for i, (h, v) in enumerate(zip(headers, row_vals)):
        print(f"Col {i} ({h}): {v}")
